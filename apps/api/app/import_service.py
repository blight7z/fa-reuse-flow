from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.orm import Session

from .errors import api_error
from .models import PartItem

EXPECTED_HEADERS = [
    "seller_ref",
    "brand",
    "model",
    "category",
    "quantity",
    "claimed_condition",
    "serial_number",
    "notes",
]
REQUIRED_HEADERS = set(EXPECTED_HEADERS) - {"serial_number", "notes"}
ALLOWED_CONDITIONS = {"N", "A", "B", "C", "D", "JUNK"}


def _text(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None


def parse_xlsx(content: bytes, filename: str, db: Session) -> list[dict]:
    if Path(filename).suffix.lower() != ".xlsx":
        raise api_error(422, "INVALID_FILE_TYPE", "รองรับเฉพาะไฟล์ .xlsx")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, ValueError, OSError, KeyError, BadZipFile):
        raise api_error(422, "INVALID_XLSX", "ไม่สามารถอ่านไฟล์ Excel ได้") from None
    sheet = next(
        (candidate for candidate in workbook.worksheets if candidate.title.casefold() == "items"),
        workbook.active,
    )
    values = sheet.iter_rows(values_only=True)
    try:
        header_values = next(values)
    except StopIteration:
        raise api_error(422, "EMPTY_FILE", "ไฟล์ไม่มีข้อมูล") from None
    headers = [_text(cell) for cell in header_values]
    missing = sorted(REQUIRED_HEADERS - set(headers))
    if missing:
        raise api_error(
            422,
            "MISSING_HEADERS",
            "ไฟล์ไม่มีคอลัมน์ที่จำเป็น",
            {header: "required header is missing" for header in missing},
        )
    header_index = {header: index for index, header in enumerate(headers) if header in EXPECTED_HEADERS}
    rows: list[dict] = []
    for row_number, cells in enumerate(values, start=2):
        if all(cell is None or str(cell).strip() == "" for cell in cells):
            continue
        data = {
            header: (cells[index] if index < len(cells) else None) for header, index in header_index.items()
        }
        normalized = {name: _text(data.get(name)) for name in EXPECTED_HEADERS}
        errors: dict[str, str] = {}
        for required in REQUIRED_HEADERS - {"quantity"}:
            if not normalized.get(required):
                errors[required] = "required"
        claimed_condition = normalized.get("claimed_condition")
        if claimed_condition and claimed_condition.upper() not in ALLOWED_CONDITIONS:
            errors["claimed_condition"] = "must be one of N, A, B, C, D, JUNK"
        elif claimed_condition:
            normalized["claimed_condition"] = claimed_condition.upper()
        quantity = data.get("quantity")
        try:
            if isinstance(quantity, bool):
                raise ValueError
            quantity_number = float(quantity)
            if not quantity_number.is_integer() or quantity_number <= 0:
                raise ValueError
            quantity_value = int(quantity_number)
        except (TypeError, ValueError, OverflowError):
            quantity_value = None
            errors["quantity"] = "must be a positive whole number"
        normalized["quantity"] = quantity_value
        rows.append(
            {
                "row_number": row_number,
                "data": normalized,
                "field_errors": errors,
            }
        )
    if not rows:
        raise api_error(422, "EMPTY_FILE", "ไฟล์ไม่มีแถวข้อมูล")

    seller_refs = {row["data"]["seller_ref"] for row in rows if row["data"]["seller_ref"]}
    if len(seller_refs) > 1:
        for row in rows:
            row["field_errors"]["seller_ref"] = "all rows in one import must use the same seller_ref"

    serial_to_rows: dict[str, list[dict]] = {}
    for row in rows:
        serial = row["data"].get("serial_number")
        if serial:
            serial_to_rows.setdefault(serial.casefold(), []).append(row)
    for duplicates in serial_to_rows.values():
        if len(duplicates) > 1:
            for row in duplicates:
                row["field_errors"]["serial_number"] = "duplicate serial in file"
    candidate_serials = [row["data"]["serial_number"] for row in rows if row["data"].get("serial_number")]
    if candidate_serials:
        existing = set(
            db.scalars(
                select(PartItem.serial_number).where(PartItem.serial_number.in_(candidate_serials))
            ).all()
        )
        for row in rows:
            if row["data"].get("serial_number") in existing:
                row["field_errors"]["serial_number"] = "serial already exists"
    for row in rows:
        row["is_valid"] = not row["field_errors"]
    return rows
